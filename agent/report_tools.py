"""
Report generation tools - Excel (.xlsx) and CSV (.csv)
"""

import os
import csv
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORTS_DIR = Path(os.environ.get("REPORTS_DIR", "/tmp/billing_reports"))
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Color palette
HEADER_FILL   = PatternFill("solid", fgColor="1A3C6E")
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E6DB4")
ANOMALY_FILL  = PatternFill("solid", fgColor="C0392B")
WARNING_FILL  = PatternFill("solid", fgColor="E67E22")
ALT_ROW_FILL  = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Calibri", bold=True, color="1A3C6E", size=14)
BODY_FONT     = Font(name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
THIN_BORDER   = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _write_header_row(ws, headers, row, fill=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = fill or HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER


def _write_data_rows(ws, rows, start_row, anomaly_col=None):
    for r_idx, row_data in enumerate(rows):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=val)
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = LEFT
            fill = ALT_ROW_FILL if r_idx % 2 == 0 else WHITE_FILL
            if anomaly_col and c_idx == 1:
                pct = row_data[anomaly_col - 1] if anomaly_col <= len(row_data) else 0
                if isinstance(pct, (int, float)):
                    fill = ANOMALY_FILL if pct >= 50 else (WARNING_FILL if pct >= 20 else fill)
            cell.fill = fill


# ---------------------------------------------------------------------------
# Tool: Generate Excel report
# ---------------------------------------------------------------------------

def generate_excel_report(
    filename: str,
    org_summary: str = "",
    project_detail: str = "",
    service_breakdown: str = "",
    anomalies: str = "",
    top_drivers: str = "",
) -> dict:
    """
    Generate a multi-sheet Excel billing report.

    Args:
        filename:          Output filename e.g. billing_report_20250614.xlsx
        org_summary:       JSON string from fetch_org_billing_summary
        project_detail:    JSON string from fetch_project_billing_detail
        service_breakdown: JSON string from fetch_service_cost_breakdown
        anomalies:         JSON string from detect_billing_anomalies
        top_drivers:       JSON string from fetch_top_cost_drivers

    Returns:
        dict with filepath, filename, sheets_created
    """
    import json

    # Parse JSON strings to dicts
    def _parse(s):
        if not s:
            return None
        try:
            return json.loads(s) if isinstance(s, str) else s
        except Exception:
            return None

    org      = _parse(org_summary)
    detail   = _parse(project_detail)
    services = _parse(service_breakdown)
    anom     = _parse(anomalies)
    drivers  = _parse(top_drivers)

    filepath = REPORTS_DIR / filename
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheets_created = []

    # ── Sheet 1: Executive Summary ───────────────────────────────────────────
    ws = wb.create_sheet("Executive Summary")
    sheets_created.append("Executive Summary")
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = f"GCP Billing Report - {datetime.now().strftime('%B %Y')}"
    title_cell.font      = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    if org:
        ws.cell(row=3, column=1, value="Report Period").font = Font(bold=True)
        ws.cell(row=3, column=2, value=f"{org.get('period', {}).get('start')} to {org.get('period', {}).get('end')}")
        ws.cell(row=4, column=1, value="Total Organization Cost").font = Font(bold=True)
        ws.cell(row=4, column=2, value=f"{org.get('currency')} {org.get('total_cost', 0):,.2f}")
        ws.cell(row=5, column=1, value="Projects Billed").font = Font(bold=True)
        ws.cell(row=5, column=2, value=org.get("project_count", 0))
        if anom:
            ws.cell(row=6, column=1, value="Anomalies Detected").font = Font(bold=True, color="C0392B")
            ws.cell(row=6, column=2, value=anom.get("summary", ""))

        headers = ["Project ID", "Project Name", "Total Cost", "Currency", "Period Start", "Period End"]
        _write_header_row(ws, headers, row=9)
        rows_data = [
            [p.get("project_id"), p.get("project_name"), p.get("total_cost"),
             p.get("currency"), str(p.get("period_start", "")), str(p.get("period_end", ""))]
            for p in org.get("projects", [])
        ]
        _write_data_rows(ws, rows_data, start_row=10)
    _auto_width(ws)

    # ── Sheet 2: Project Detail ──────────────────────────────────────────────
    if detail and detail.get("rows"):
        ws2 = wb.create_sheet("Project Detail")
        sheets_created.append("Project Detail")
        headers = ["Project ID", "Project Name", "Date", "Service", "Daily Cost", "Currency"]
        _write_header_row(ws2, headers, row=1)
        rows_data = [
            [r.get("project_id"), r.get("project_name"), str(r.get("usage_date")),
             r.get("service"), r.get("daily_cost"), r.get("currency")]
            for r in detail["rows"]
        ]
        _write_data_rows(ws2, rows_data, start_row=2)
        _auto_width(ws2)

    # ── Sheet 3: Service Breakdown ───────────────────────────────────────────
    if services and services.get("services"):
        ws3 = wb.create_sheet("Service Breakdown")
        sheets_created.append("Service Breakdown")
        headers = ["Service ID", "Service Name", "Total Cost", "Currency", "Projects Using"]
        _write_header_row(ws3, headers, row=1, fill=SUBHEAD_FILL)
        rows_data = [
            [s.get("service_id"), s.get("service_name"), s.get("total_cost"),
             s.get("currency"), s.get("project_count")]
            for s in services["services"]
        ]
        _write_data_rows(ws3, rows_data, start_row=2)
        _auto_width(ws3)

    # ── Sheet 4: Anomalies ───────────────────────────────────────────────────
    if anom and anom.get("anomalies"):
        ws4 = wb.create_sheet("Anomalies")
        sheets_created.append("Anomalies")
        headers = ["Project ID", "Project Name", "Service", "Date",
                   "Daily Cost", "7-Day Avg", "% Change", "Currency"]
        _write_header_row(ws4, headers, row=1, fill=ANOMALY_FILL)
        rows_data = [
            [a.get("project_id"), a.get("project_name"), a.get("service"),
             str(a.get("usage_date")), a.get("daily_cost"), a.get("avg_7d_cost"),
             a.get("pct_change"), a.get("currency")]
            for a in anom["anomalies"]
        ]
        _write_data_rows(ws4, rows_data, start_row=2, anomaly_col=7)
        _auto_width(ws4)

    # ── Sheet 5: Top Cost Drivers ────────────────────────────────────────────
    if drivers and drivers.get("drivers"):
        ws5 = wb.create_sheet("Top Cost Drivers")
        sheets_created.append("Top Cost Drivers")
        headers = ["Rank", "Project ID", "Project Name", "Service", "SKU", "Total Cost", "Currency"]
        _write_header_row(ws5, headers, row=1, fill=SUBHEAD_FILL)
        rows_data = [
            [i+1, d.get("project_id"), d.get("project_name"), d.get("service"),
             d.get("sku"), d.get("total_cost"), d.get("currency")]
            for i, d in enumerate(drivers["drivers"])
        ]
        _write_data_rows(ws5, rows_data, start_row=2)
        _auto_width(ws5)

    wb.save(filepath)
    return {
        "filepath":       str(filepath),
        "filename":       filename,
        "sheets_created": sheets_created,
        "size_bytes":     filepath.stat().st_size,
    }


# ---------------------------------------------------------------------------
# Tool: Generate CSV report
# ---------------------------------------------------------------------------

def generate_csv_report(
    filename: str,
    org_summary: str = "",
    anomalies: str = "",
    top_drivers: str = "",
) -> dict:
    """
    Generate a flat CSV billing report with summary, anomalies and top drivers.

    Args:
        filename:     Output filename e.g. billing_report_20250614.csv
        org_summary:  JSON string from fetch_org_billing_summary
        anomalies:    JSON string from detect_billing_anomalies
        top_drivers:  JSON string from fetch_top_cost_drivers

    Returns:
        dict with filepath, filename, row_count
    """
    import json

    def _parse(s):
        if not s:
            return None
        try:
            return json.loads(s) if isinstance(s, str) else s
        except Exception:
            return None

    org     = _parse(org_summary)
    anom    = _parse(anomalies)
    drivers = _parse(top_drivers)

    filepath  = REPORTS_DIR / filename
    row_count = 0

    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)

        writer.writerow(["# ORGANIZATION BILLING SUMMARY"])
        if org:
            writer.writerow(["Period Start", "Period End", "Total Cost", "Currency", "Project Count"])
            writer.writerow([
                org.get("period", {}).get("start"),
                org.get("period", {}).get("end"),
                org.get("total_cost"),
                org.get("currency"),
                org.get("project_count"),
            ])
            row_count += 1
            writer.writerow([])
            writer.writerow(["# PROJECT BREAKDOWN"])
            writer.writerow(["Project ID", "Project Name", "Total Cost", "Currency"])
            for p in org.get("projects", []):
                writer.writerow([p.get("project_id"), p.get("project_name"),
                                 p.get("total_cost"), p.get("currency")])
                row_count += 1
            writer.writerow([])

        writer.writerow(["# BILLING ANOMALIES"])
        if anom and anom.get("anomalies"):
            writer.writerow(["Project ID", "Project Name", "Service", "Date",
                             "Daily Cost", "7-Day Avg", "% Change", "Currency", "Severity"])
            for a in anom["anomalies"]:
                severity = "CRITICAL" if a.get("pct_change", 0) >= 50 else "WARNING"
                writer.writerow([
                    a.get("project_id"), a.get("project_name"), a.get("service"),
                    a.get("usage_date"), a.get("daily_cost"), a.get("avg_7d_cost"),
                    a.get("pct_change"), a.get("currency"), severity,
                ])
                row_count += 1
        else:
            writer.writerow(["No anomalies detected"])
        writer.writerow([])

        writer.writerow(["# TOP COST DRIVERS"])
        if drivers and drivers.get("drivers"):
            writer.writerow(["Rank", "Project ID", "Project Name", "Service", "SKU", "Total Cost", "Currency"])
            for i, d in enumerate(drivers["drivers"]):
                writer.writerow([i+1, d.get("project_id"), d.get("project_name"),
                                 d.get("service"), d.get("sku"), d.get("total_cost"), d.get("currency")])
                row_count += 1

    return {
        "filepath":   str(filepath),
        "filename":   filename,
        "row_count":  row_count,
        "size_bytes": filepath.stat().st_size,
    }